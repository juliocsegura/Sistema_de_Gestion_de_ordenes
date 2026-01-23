from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
import openpyxl
import pandas as pd
from datetime import datetime, time
from django.contrib.contenttypes.models import ContentType
from django.db.models import Q, Prefetch
from django.db import transaction, models
from .models import (Moldmakers, OrdenMCM, OrdenCHO, OrdenTPM, OrdenPREP,EstatusOrden,Maquinas,ActividadTPM,ActividadPREP,SubZonaTPM,ZonaTPM,NumerosParte, Moldes, OrdenSAP, Defectos, AsignacionUniversal,Lideres)
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.http import require_http_methods 
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import json
from itertools import chain, zip_longest
from operator import attrgetter
def es_lider(user):
    if not user.is_authenticated: 
        return False
    return user.is_superuser or user.groups.filter(name='Lideres').exists()

# --- VISTA DEL PANEL PRINCIPAL (Dashboard) ---
def panel_view(request):
    return render(request, 'Moldeo/panel_principal.html')

# --- VISTAS DE REGISTRO ---

def Registrar_Orden_view(request):
    lista_tecnicos = Moldmakers.objects.all().order_by('nombre')
    es_lider_user = request.user.groups.filter(name='Lideres').exists() or request.user.is_superuser
    
    # Datos de URL
    orden_sap = request.GET.get('orden_sap', '') or request.GET.get('numero_orden', '')
    defecto = request.GET.get('defecto', '') or request.GET.get('defecto_sap', '')
    molde_str = request.GET.get('molde', '') # Nombre (Ej: M123)
    status_url = request.GET.get('status', '100')
    maquina = request.POST.get('maquina','') or request.GET.get('maquina', '')
    # BUSCAR ID DEL MOLDE (Crucial)
    pre_molde_pk = ''
    if molde_str:
        try:
            m = Moldes.objects.filter(nombre=molde_str).first()
            if m: pre_molde_pk = m.id_molde
        except: pass

    ctx = {
        'pre_orden': orden_sap,
        'pre_defecto': defecto,
        'pre_maquina': maquina,
        'pre_molde_nombre': molde_str,
        'pre_molde_pk': pre_molde_pk, # <--- Esto llena el input hidden
        'status_actual': status_url,
        'moldmakers': lista_tecnicos,
        'es_lider_logueado': es_lider_user,
        'lideres_all': Lideres.objects.all().order_by('nombre'),
        'lista_defectos': Defectos.objects.all().order_by('nombre_defecto')
    }
    return render(request, 'Moldeo/registrar_orden.html', ctx)

def Orden_en_curso_view(request):
    lista_tecnicos = Moldmakers.objects.all().order_by('nombre')
    es_lider = request.user.is_superuser or request.user.groups.filter(name='Lideres').exists()
    lista_defectos = Defectos.objects.all().order_by('nombre_defecto')
    # DEBUG: Imprimir en la terminal de Python para verificar
    print(f"DEBUG PANEL: Usuario {request.user.username} - SuperUser: {request.user.is_superuser} - Es Lider: {es_lider}")
    context = { 
        'lista_defectos': lista_defectos,
        'moldmakers': lista_tecnicos,
        'lideres_all': Lideres.objects.all().order_by('nombre'),
    }
    return render(request, 'Moldeo/Ordenes_en_curso.html', context)

def btn_status_ordenmcm_view(request):
    context = {
        'pre_orden': request.GET.get('numero_orden', ''),
        'pre_defecto': request.GET.get('defecto_sap', ''),
        'pre_molde': request.GET.get('molde', ''),
        
        'pre_maquina': request.GET.get('maquina', '')
    }
    return render(request, 'Moldeo/btn_status_mcm.html', context)
def btn_status_ordentpm_view(request):
    context = {
        'pre_orden': request.GET.get('numero_orden', ''),
        'pre_defecto': request.GET.get('defecto_sap', ''),
        'pre_molde': request.GET.get('molde', ''),
        
        'pre_maquina': request.GET.get('maquina', '')
    }
    return render(request, 'Moldeo/btn_status_tpm.html', context)
def btn_status_ordenprep_view(request):
    context = {
        'pre_orden': request.GET.get('numero_orden', ''),
        'pre_defecto': request.GET.get('defecto_sap', ''),
        'pre_molde': request.GET.get('molde', ''),
        
        'pre_maquina': request.GET.get('maquina', '')
    }
    return render(request, 'Moldeo/btn_status_prep.html', context)

@require_http_methods(["GET", "POST"])
def mcm_view(request):
    # 1. RECUPERAR PARÁMETROS INICIALES (GET o POST)
    status_actual = request.POST.get('statusmcm') or request.GET.get('status', '')
    pre_orden = request.POST.get('numero_orden') or request.GET.get('numero_orden', '')
    pre_defecto = request.POST.get('defecto_sap') or request.GET.get('defecto_sap', '')
    pre_maquina = request.POST.get('maquina') or request.GET.get('maquina', '')
    
    # Manejo inteligente del Molde (Nombre vs ID)
    pre_molde_nombre = request.GET.get('molde', '') 
    pre_molde_pk = request.POST.get('molde', '')

    molde_obj = None

    # Caso A: Tenemos ID (POST), buscamos el objeto
    if pre_molde_pk:
        try:
            molde_obj = Moldes.objects.get(pk=pre_molde_pk)
            pre_molde_nombre = molde_obj.nombre
            if not pre_maquina and molde_obj.maquina: # Autocompletar máquina si falta
                pre_maquina = molde_obj.maquina.nombre
        except: pass
    
    # Caso B: Tenemos Nombre (GET), buscamos ID y Máquina
    elif pre_molde_nombre:
        try:
            molde_obj = Moldes.objects.filter(nombre=pre_molde_nombre).first()
            if molde_obj: 
                pre_molde_pk = molde_obj.pk
                if not pre_maquina and molde_obj.maquina:
                    pre_maquina = molde_obj.maquina.nombre
        except: pass

    # 2. LÓGICA DE GUARDADO (POST)
    if request.method == 'POST':
        # Recuperamos datos del formulario
        numero_orden = request.POST.get('numero_orden')
        defecto_sap = request.POST.get('defecto_sap')
        molde_id = request.POST.get('molde')
        asignaciones_json = request.POST.get('asignaciones_json', '[]')
        texto_auto = request.POST.get('comentarios_generados', '')
        # Recuperar máquina (del input o del objeto molde)
        maquina_final = request.POST.get('maquina')
        if not maquina_final and molde_obj:
             maquina_final = molde_obj.maquina.nombre if molde_obj.maquina else ''

        # Recuperar datos de retorno
        motivo_ret = request.POST.get('motivo_retorno', '')
        obs_ret = request.POST.get('observaciones_retorno', '')
        orden_ref = request.POST.get('orden_retorno_ref', '')
        obs_manual = request.POST.get('observaciones_retorno', '') 
        comentarios_finales = texto_auto
        if obs_manual:
            comentarios_finales += f"\n\n[Notas Adicionales]:\n{obs_manual}"
        # Validaciones
        if not numero_orden:
            messages.error(request, 'Error: Falta el Número de Orden.')
        elif not molde_obj:
            messages.error(request, 'Error: El Molde no es válido o no existe.')
        else:
            try:
                with transaction.atomic():
                    # Crear la Orden
                    nueva_orden = OrdenMCM.objects.create(
                        numero_orden=numero_orden,
                        defecto_sap=defecto_sap,
                        molde=molde_obj,
                        status=status_actual,
                        tipo_mntn='MCM',
                        maquina=maquina_final, # Usamos la máquina recuperada
                        estado='Activa',
                        ultima_actualizacion=timezone.now(),
                        motivo_retorno=motivo_ret,
                        comentarios=comentarios_finales,
                        orden_retorno_ref=orden_ref,
                        observaciones_retorno=obs_ret
                    )

                    # Procesar Técnicos (JSON)
                    asignaciones_data = json.loads(asignaciones_json)
                    
                    for item in asignaciones_data:
                        nombre = item.get('nombre')
                        lider_tec = item.get('lider', '')
                        mesa = item.get('mesa', '-')
                        
                        if nombre:
                            # Formato nombre: "Juan Perez (L: Carlos)"
                            nombre_guardar = f"{nombre} (L: {lider_tec})" if lider_tec else nombre
                            detalles_str = json.dumps(item.get('detalles', []))

                            AsignacionUniversal.objects.create(
                                content_object=nueva_orden,
                                nombre_tecnico=nombre_guardar,
                                mesa=mesa,
                                detalles_json=detalles_str,
                                defecto="Ver detalles",
                                activo=True
                            )

                messages.success(request, f'Orden MCM {numero_orden} registrada correctamente.')
                return redirect('Moldeo:ordenes_en_curso')

            except Exception as e:
                messages.error(request, f'Error crítico al guardar: {str(e)}')
                print(f"DEBUG ERROR SAVE: {e}")

    # 3. RENDERIZADO (GET o Fallo POST)
    context = {
        'status_actual': status_actual,
        'tipo_mntn': 'MCM',
        'pre_orden': pre_orden,
        'pre_defecto': pre_defecto,
        'pre_maquina': pre_maquina, # Pasamos la máquina recuperada
        'pre_molde_nombre': pre_molde_nombre, 
        'pre_molde_pk': pre_molde_pk,
        'lista_defectos': Defectos.objects.filter(activo=True).order_by('nombre_defecto'),
        'moldmakers': Moldmakers.objects.filter(activo=True).order_by('nombre'),
        'lideres_all': Lideres.objects.filter(activo=True).order_by('nombre'),
    }
    return render(request, 'Moldeo/prueba.html', context)
@require_http_methods(["GET", "POST"])
@transaction.atomic
def registro_cho_view(request):
    # 1. Configuración Inicial y Recuperación de Datos
    status_actual = '110' # Valor por defecto para CHO
    tipo_mntn = 'CHO'
    
    pre_orden = request.POST.get('numero_orden') or request.GET.get('numero_orden', '')
    pre_defecto = request.POST.get('defecto_sap') or request.GET.get('defecto_sap', '')
    pre_maquina = request.POST.get('maquina') or request.GET.get('maquina', '')
    
    # Manejo inteligente del Molde (Nombre vs ID)
    pre_molde_nombre = request.GET.get('molde', '') 
    pre_molde_pk = request.POST.get('molde', '')
    
    molde_obj = None
    partes_data = [] # Lista para el dropdown

    # Búsqueda inteligente del Molde
    if pre_molde_pk:
        try:
            molde_obj = Moldes.objects.get(pk=pre_molde_pk)
            pre_molde_nombre = molde_obj.nombre
        except: pass
    elif pre_molde_nombre:
        try:
            molde_obj = Moldes.objects.filter(nombre=pre_molde_nombre).first()
            if molde_obj: pre_molde_pk = molde_obj.pk
        except: pass

    # Obtener Números de Parte si tenemos molde
    if molde_obj:
        qs_partes = NumerosParte.objects.filter(molde=molde_obj).values_list('numero_parte', flat=True)
        partes_data = [{'nombre': p} for p in qs_partes]
    
    partes_json = json.dumps(partes_data)

    # 2. PROCESAMIENTO POST
    if request.method == 'POST':
        # --- DATOS GENERALES ---
        numero_orden = request.POST.get('numero_orden')
        defecto_sap = request.POST.get('defecto_sap')
        
        # Recuperar máquina (del input o del objeto)
        maquina_input = request.POST.get('maquina', '')
        if not maquina_input and molde_obj and molde_obj.maquina:
            maquina_input = molde_obj.maquina.nombre

        # Datos específicos de CHO (Orden Principal)
        parte_saliente = request.POST.get('parte_saliente', '')
        parte_entrante = request.POST.get('parte_entrante', '')

        # Datos específicos de Tarjeta Roja (Orden Vinculada)
        tipo_tarjeta = request.POST.get('tipo_tarjeta', 'verde')
        orden_vinculada_id = request.POST.get('orden_vinculada_id', '')
        parte_saliente_vinc = request.POST.get('parte_saliente_vinc', '')
        parte_entrante_vinc = request.POST.get('parte_entrante_vinc', '')
        copiar_tecnicos = request.POST.get('copiar_tecnicos') == 'true'

        asignaciones_json = request.POST.get('asignaciones_json', '')
        comentarios_generados = request.POST.get('comentarios_generados', '')
        
        if not (numero_orden and molde_obj):
            messages.error(request, 'Error: Faltan campos obligatorios (Orden o Molde).')
        else:
            try:
                # =================================================
                # A. GUARDAR ORDEN PRINCIPAL
                # =================================================
                nueva_orden = OrdenCHO.objects.create(
                    numero_orden=numero_orden,
                    defecto_sap=defecto_sap,
                    molde=molde_obj,
                    tipo_mntn=tipo_mntn,
                    status=status_actual,
                    estado='Activa',
                    maquina=maquina_input,
                    parte_saliente=parte_saliente, 
                    parte_entrante=parte_entrante, 
                    # tipo_tarjeta=tipo_tarjeta, # Descomentar si agregaste el campo al modelo
                    comentarios=comentarios_generados,
                    ultima_actualizacion=timezone.now()
                )

                # Guardar Técnicos Principal
                if asignaciones_json:
                    data = json.loads(asignaciones_json)
                    for item in data:
                        nombre = item.get('nombre')
                        lider_tec = item.get('lider', '')
                        mesa = item.get('mesa', '-')
                        
                        if nombre:
                            nombre_final = f"{nombre} (L: {lider_tec})" if lider_tec else nombre
                            detalles_str = json.dumps(item.get('detalles', []))
                            
                            AsignacionUniversal.objects.create(
                                content_object=nueva_orden,
                                nombre_tecnico=nombre_final,
                                mesa=mesa,
                                detalles_json=detalles_str,
                                activo=True
                            )
                else:
                    # Fallback para arrays simples (si aplica)
                    l_tecnicos = request.POST.getlist('tecnico_nombre[]')
                    l_mesas = request.POST.getlist('mesa[]')
                    from itertools import zip_longest
                    for tec, mes in zip_longest(l_tecnicos, l_mesas, fillvalue=''):
                        if tec and tec.strip():
                            AsignacionUniversal.objects.create(
                                content_object=nueva_orden,
                                nombre_tecnico=tec.strip(),
                                mesa=mes.strip(),
                                activo=True
                            )

                # =================================================
                # B. PROCESAR ORDEN VINCULADA (SOLO SI ES ROJA)
                # =================================================
                if tipo_tarjeta == 'roja' and orden_vinculada_id:
                    # Buscar la orden existente que estaba pendiente
                    orden_vinc = OrdenCHO.objects.filter(numero_orden=orden_vinculada_id).first()
                    
                    if orden_vinc:
                        # Actualizar datos para iniciarla
                        orden_vinc.estado = 'Activa'
                        orden_vinc.status = status_actual
                        orden_vinc.maquina = maquina_input
                        orden_vinc.molde = molde_obj 
                        
                        # Usar sus propios datos de NP (Independientes)
                        orden_vinc.parte_saliente = parte_saliente_vinc
                        orden_vinc.parte_entrante = parte_entrante_vinc
                        
                        # Comentario de enlace
                        comentario_extra = f"\n[Iniciada vía Tarjeta Roja vinculada a {numero_orden}]"
                        if orden_vinc.comentarios:
                            orden_vinc.comentarios += comentario_extra
                        else:
                            orden_vinc.comentarios = comentario_extra
                            
                        orden_vinc.ultima_actualizacion = timezone.now()
                        orden_vinc.save()

                        # Copiar técnicos si se solicitó
                        if copiar_tecnicos and asignaciones_json:
                            for item in data:
                                nombre = item.get('nombre')
                                lider_tec = item.get('lider', '')
                                mesa = item.get('mesa', '-')
                                if nombre:
                                    nombre_final = f"{nombre} (L: {lider_tec})" if lider_tec else nombre
                                    AsignacionUniversal.objects.create(
                                        content_object=orden_vinc,
                                        nombre_tecnico=nombre_final,
                                        mesa=mesa,
                                        detalles_json="[]",
                                        activo=True
                                    )
                        
                        messages.success(request, f'¡Éxito! Orden principal {numero_orden} y vinculada {orden_vinculada_id} iniciadas.')
                    else:
                        messages.warning(request, f'Orden principal guardada, pero no se encontró la orden vinculada {orden_vinculada_id}.')
                else:
                    messages.success(request, f'Orden CHO {numero_orden} registrada.')

                return redirect('Moldeo:ordenes_en_curso')

            except Exception as e:
                messages.error(request, f'Error al guardar: {e}')
                print(f"Error CHO: {e}")

    # 3. RENDERIZADO
    context = {
        'tipo_mntn': tipo_mntn,
        'status_actual': status_actual,
        'pre_orden': pre_orden,
        'pre_molde_nombre': pre_molde_nombre,
        'pre_molde_pk': pre_molde_pk,
        'pre_defecto': pre_defecto,
        'pre_maquina': molde_obj.maquina.nombre if molde_obj and molde_obj.maquina else '',
        'moldmakers': Moldmakers.objects.filter(activo=True).order_by('nombre'),
        'lideres_all': Lideres.objects.filter(activo=True).order_by('nombre'),
        'partes_json': partes_json 
    }
    return render(request, 'Moldeo/registro_cho.html', context)
@require_http_methods(["GET", "POST"])
@transaction.atomic
def registro_tpm_view(request):
    # 1. Configuración Inicial
    tipo_mntn = 'TPM'
    status_actual = '206' # O el estatus default para TPM

    # Recuperar datos de GET (para precargar)
    pre_orden = request.GET.get('numero_orden', '') or request.POST.get('numero_orden', '')
    pre_maquina = request.POST.get('maquina') or request.GET.get('maquina', '')
    pre_defecto = request.POST.get('defecto_sap') or request.GET.get('defecto_sap', '')
    # Manejo inteligente del Molde (Nombre vs ID)
    pre_molde_nombre = request.GET.get('molde', '') 
    pre_molde_pk = request.POST.get('molde', '')
    molde_obj = None

    # Búsqueda inteligente del Molde (Igual que en CHO)
    if pre_molde_pk:
        try:
            molde_obj = Moldes.objects.get(pk=pre_molde_pk)
            pre_molde_nombre = molde_obj.nombre
        except: pass
    elif pre_molde_nombre:
        try:
            molde_obj = Moldes.objects.filter(nombre=pre_molde_nombre).first()
            if molde_obj: pre_molde_pk = molde_obj.pk
        except: pass

    # ==============================================================
    # 2. PREPARACIÓN DE DATOS (CATÁLOGOS PARA JS)
    # ==============================================================
    
    # A. Actividades
    actividades = list(ActividadTPM.objects.filter(activo=True).values('nombre'))
    
    # B. Zonas y Subzonas (Estructura Anidada para el JS)
    # Esto genera: [{nombre: 'MODULOS', subzonas: [{nombre: 'General', req: false}, ...]}, ...]
    zonas_qs = ZonaTPM.objects.filter(activo=True).prefetch_related('subzonas')
    zonas_data = []
    
    for z in zonas_qs:
        subs = []
        for s in z.subzonas.filter(activo=True):
            # 'req' le dice al JS si debe mostrar los inputs de modulo/cavidad/lado
            subs.append({
                'nombre': s.nombre,
                'req': str(s.requiere_detalles).lower() # 'true'/'false' para JS
            })
        zonas_data.append({
            'nombre': z.nombre,
            'subzonas': subs
        })

    # Convertir a JSON strings seguros para el template
    actividades_json = json.dumps(actividades)
    zonas_json = json.dumps(zonas_data)
    tecnicos_json = json.dumps([{'nombre': t.nombre} for t in Moldmakers.objects.filter(activo=True).order_by('nombre')])
    lideres_json = json.dumps([{'nombre': l.nombre} for l in Lideres.objects.filter(activo=True).order_by('nombre')])

    # ==============================================================
    # 3. PROCESAMIENTO POST (GUARDADO)
    # ==============================================================
    if request.method == 'POST':
        numero_orden = request.POST.get('numero_orden')
        # Intentamos recuperar el molde del POST si no vino antes
        if not molde_obj and request.POST.get('molde'):
             try:
                 molde_obj = Moldes.objects.get(pk=request.POST.get('molde'))
             except: pass

        # Recuperar el JSON con toda la estructura (Técnicos + Actividades)
        asignaciones_json = request.POST.get('asignaciones_json', '')

        if not (numero_orden and molde_obj):
            messages.error(request, 'Error: Faltan datos obligatorios (Orden o Molde).')
        else:
            try:
                # Crear la Orden TPM
                nueva_orden = OrdenTPM.objects.create(
                    numero_orden=numero_orden,
                    tipo_mntn = tipo_mntn,
                    molde=molde_obj,
                    maquina=molde_obj.maquina.nombre if molde_obj.maquina else None,
                    status=status_actual, # Status numérico ej. 110
                    estado='Activa',      # Estado legible ej. Activa
                    comentarios="Inicio de Mantenimiento TPM",
                    ultima_actualizacion=timezone.now()
                )

                # Procesar Asignaciones (JSON Complejo)
                if asignaciones_json:
                    data = json.loads(asignaciones_json)
                    
                    for item in data:
                        nombre = item.get('nombre')
                        lider_tec = item.get('lider', '')
                        mesa = item.get('mesa', '-')
                        lista_actividades = item.get('actividades', [])

                        if nombre:
                            nombre_final = f"{nombre} (L: {lider_tec})" if lider_tec else nombre
                            
                            detalles_str = json.dumps(lista_actividades)

                            AsignacionUniversal.objects.create(
                                content_object=nueva_orden,
                                nombre_tecnico=nombre_final,
                                mesa=mesa,
                                detalles_json=detalles_str, 
                                activo=True
                            )

                messages.success(request, f'Orden TPM {numero_orden} iniciada correctamente.')
                return redirect('Moldeo:ordenes_en_curso')

            except Exception as e:
                messages.error(request, f'Error al guardar TPM: {str(e)}')
                print(f"Error TPM View: {e}") # Debug en consola

    # ==============================================================
    # 4. RENDERIZADO
    # ==============================================================
    context = {
        'tipo_mntn': tipo_mntn,
        'pre_orden': pre_orden,
        'pre_molde_nombre': pre_molde_nombre,
        'pre_molde_pk': pre_molde_pk,
        'pre_maquina': molde_obj.maquina.nombre if molde_obj and molde_obj.maquina else '',
        # Pasamos los JSONs para que los use el JavaScript
        'actividades_json': actividades_json,
        'zonas_json': zonas_json,
        'pre_defecto': pre_defecto,
        'tecnicos_json': tecnicos_json,
        'lideres_json': lideres_json
    }
    return render(request, 'Moldeo/registro_tpm.html', context)
@require_http_methods(["GET", "POST"])
@transaction.atomic
def registro_prep_view(request):
    # 1. Configuración Inicial
    tipo_mntn = 'PREP'
    status_actual = request.POST.get('statusmcm') or request.GET.get('status', '')

    # Recuperar datos de GET/POST (para precargar)
    pre_orden = request.GET.get('numero_orden', '') or request.POST.get('numero_orden', '')
    pre_maquina = request.POST.get('maquina') or request.GET.get('maquina', '')
    pre_defecto = request.POST.get('defecto_sap') or request.GET.get('defecto_sap', '')
    # Manejo inteligente del Molde (Nombre vs ID)
    pre_molde_nombre = request.GET.get('molde', '') 
    pre_molde_pk = request.POST.get('molde', '')
    molde_obj = None

    # Búsqueda inteligente del Molde
    if pre_molde_pk:
        try:
            molde_obj = Moldes.objects.get(pk=pre_molde_pk)
            pre_molde_nombre = molde_obj.nombre
        except: pass
    elif pre_molde_nombre:
        try:
            molde_obj = Moldes.objects.filter(nombre=pre_molde_nombre).first()
            if molde_obj: pre_molde_pk = molde_obj.pk
        except: pass

    # ==============================================================
    # 2. PREPARACIÓN DE DATOS (CATÁLOGOS PARA JS)
    # ==============================================================
    
    # A. Actividades PREP
    actividades = list(ActividadPREP.objects.filter(activo=True).values('nombre'))
    
    # B. Zonas (Vacío para PREP)
    zonas_data = [] 
    
    # Convertir a JSON strings seguros para el template
    actividades_json = json.dumps(actividades)
    zonas_json = json.dumps(zonas_data) 
    tecnicos_json = json.dumps([{'nombre': t.nombre} for t in Moldmakers.objects.filter(activo=True).order_by('nombre')])
    lideres_json = json.dumps([{'nombre': l.nombre} for l in Lideres.objects.filter(activo=True).order_by('nombre')])

    # ==============================================================
    # 3. PROCESAMIENTO POST (GUARDADO)
    # ==============================================================
    if request.method == 'POST':
        numero_orden = request.POST.get('numero_orden')
        # Intentamos recuperar el molde del POST si no vino antes
        if not molde_obj and request.POST.get('molde'):
             try:
                 molde_obj = Moldes.objects.get(pk=request.POST.get('molde'))
             except: pass

        asignaciones_json = request.POST.get('asignaciones_json', '')

        if not (numero_orden and molde_obj):
            messages.error(request, 'Error: Faltan datos obligatorios (Orden o Molde).')
        else:
            try:
                # Crear la Orden PREP
                nueva_orden = OrdenPREP.objects.create(
                    numero_orden=numero_orden,
                    tipo_mntn = tipo_mntn,
                    molde=molde_obj,
                    maquina=molde_obj.maquina.nombre if molde_obj.maquina else None,
                    status=status_actual, # Valor fijo
                    estado='Activa',      # Estado legible
                    comentarios="Inicio de Mantenimiento PREP",
                    ultima_actualizacion=timezone.now()
                )

                # Procesar Asignaciones (JSON)
                if asignaciones_json:
                    data = json.loads(asignaciones_json)
                    
                    for item in data:
                        nombre = item.get('nombre')
                        lider_tec = item.get('lider', '')
                        mesa = item.get('mesa', '-')
                        lista_actividades = item.get('actividades', [])

                        if nombre:
                            nombre_final = f"{nombre} (L: {lider_tec})" if lider_tec else nombre
                            
                            # Guardamos las actividades en el JSON 'detalles_json'
                            detalles_str = json.dumps(lista_actividades)

                            AsignacionUniversal.objects.create(
                                content_object=nueva_orden,
                                nombre_tecnico=nombre_final,
                                mesa=mesa,
                                detalles_json=detalles_str, 
                                activo=True
                            )

                messages.success(request, f'Orden PREP {numero_orden} iniciada correctamente.')
                return redirect('Moldeo:ordenes_en_curso')

            except Exception as e:
                messages.error(request, f'Error al guardar PREP: {str(e)}')
                print(f"Error PREP View: {e}") # Debug en consola

    # ==============================================================
    # 4. RENDERIZADO
    # ==============================================================
    context = {
        'status_actual': status_actual,      
        'tipo_mntn': tipo_mntn,
        'pre_orden': pre_orden,
        'pre_molde_nombre': pre_molde_nombre,
        'pre_molde_pk': pre_molde_pk,
        'pre_defecto': pre_defecto,
        'pre_maquina': molde_obj.maquina.nombre if molde_obj and molde_obj.maquina else '',
        'actividades_json': actividades_json,
        'zonas_json': zonas_json,            
        'tecnicos_json': tecnicos_json,
        'lideres_json': lideres_json
    }
    return render(request, 'Moldeo/registro_prep.html', context)
# --- VISTAS DE API ---

@require_http_methods(["GET"])
def api_ordenes_recientes_view(request):
    # 1. Optimización de Consultas (Prefetch)
    p_asignaciones = Prefetch('asignaciones', queryset=AsignacionUniversal.objects.filter(activo=True))

    qs_mcm = OrdenMCM.objects.exclude(estado='Finalizada').select_related('molde', 'lider').prefetch_related(p_asignaciones)
    qs_cho = OrdenCHO.objects.exclude(estado='Finalizada').select_related('molde', 'lider').prefetch_related(p_asignaciones)
    qs_tpm = OrdenTPM.objects.exclude(estado='Finalizada').select_related('molde', 'lider').prefetch_related(p_asignaciones)
    qs_prep = OrdenPREP.objects.exclude(estado='Finalizada').select_related('molde', 'lider').prefetch_related(p_asignaciones)

    # 2. Unificar listas
    todas = list(chain(qs_mcm, qs_cho, qs_tpm, qs_prep))
    todas.sort(key=attrgetter('fecha_creacion'), reverse=True)
    
    data = []

    # 3. Bucle Único
    for orden in todas:
        tecnicos_data = []
        elementos_visuales = set() # Aquí guardaremos Defectos (MCM) o Actividades (TPM)
        nombres_tecnicos_visibles = []

        # Recorrer asignaciones activas
        asignaciones = orden.asignaciones.all() # Ya filtrado por el Prefetch a activo=True

        for asig in asignaciones:
            nombres_tecnicos_visibles.append(asig.nombre_tecnico)
            
            # Parsear JSON de detalles
            detalles_obj = []
            if asig.detalles_json:
                try:
                    detalles_obj = json.loads(asig.detalles_json)
                except: pass
            
            # --- LÓGICA DIFERENCIADA ---
            if orden.tipo_mntn in ['TPM','PREP']:
                # Si es TPM, buscamos "actividad"
                if isinstance(detalles_obj, list):
                    for item in detalles_obj:
                        if 'actividad' in item:
                            elementos_visuales.add(item['actividad'])
            else:
                # Si es CHO/MCM, buscamos "defecto"
                if isinstance(detalles_obj, list):
                    for item in detalles_obj:
                        if 'defecto' in item:
                            elementos_visuales.add(item['defecto'])
                
                # Compatibilidad con datos viejos planos
                if not detalles_obj and getattr(asig, 'defecto', None):
                     elementos_visuales.add(asig.defecto)

            # Extraer cavidad/circuito del primer detalle para la tarjeta (MCM)
            resumen_cav = '-'
            resumen_circ = '-'
            if detalles_obj and isinstance(detalles_obj, list) and len(detalles_obj) > 0:
                resumen_cav = detalles_obj[0].get('cav', detalles_obj[0].get('cavidad', '-'))
                resumen_circ = detalles_obj[0].get('circ', detalles_obj[0].get('circuito', '-'))
            elif orden.tipo_mntn == 'TPM' and detalles_obj:
                 # En TPM podemos mostrar la Zona como "cavidad" visualmente si quieres
                 resumen_cav = detalles_obj[0].get('zona', '-')

            # Construir objeto del técnico para el modal
            tecnicos_data.append({
                'id': asig.id, 
                'nombre': asig.nombre_tecnico,
                'mesa': asig.mesa or '-',
                'activo': asig.activo,
                'lider': str(orden.lider) if orden.lider else '-',
                'detalles': detalles_obj,
                'cavidad': resumen_cav,
                'circuito': resumen_circ
            })

        # --- DEFINIR TEXTO DE LA COLUMNA "DEFECTOS / ACTIVIDADES" ---
        if elementos_visuales:
            # Si encontramos actividades (TPM) o defectos (MCM) en el JSON
            texto_defectos = ", ".join(list(elementos_visuales))
        else:
            # Fallback: Usar el defecto de SAP o mensaje genérico
            if orden.tipo_mntn == 'TPM':
                texto_defectos = "Mantenimiento Preventivo"
            else:
                texto_defectos = getattr(orden, 'defecto_sap', 'Sin defecto registrado')

        # Datos del primer técnico (para mostrar en tarjeta principal)
        primero_activo = tecnicos_data[0] if tecnicos_data else None
        
        # Nombre del molde seguro
        nombre_molde = "N/A"
        if orden.molde:
            try: nombre_molde = orden.molde.nombre
            except: nombre_molde = "Ref Error"

        # --- CONSTRUCCIÓN FINAL DEL OBJETO ---
        data.append({
            'id': orden.id,
            'numero_orden': orden.numero_orden,
            'status': getattr(orden, 'status', '-'),
            'tipo': orden.tipo_mntn,
            'fecha_creacion': timezone.localtime(orden.fecha_creacion).strftime('%d/%m/%Y %H:%M'),
            'molde': nombre_molde,
            'defecto_sap': getattr(orden, 'defecto_sap', '-'),
            
            # AQUÍ ESTÁ LA MAGIA: lista_defectos contiene Actividades si es TPM
            'lista_defectos': texto_defectos, 
            
            'estado': orden.estado, 
            'comentarios': orden.comentarios, 
            'tecnico': ", ".join(nombres_tecnicos_visibles) if nombres_tecnicos_visibles else "Sin Asignar",
            'tecnicos_lista': tecnicos_data,
            'maquina': orden.maquina if orden.maquina else None,
            'mesa': primero_activo['mesa'] if primero_activo else '-',
            'cavidad': primero_activo['cavidad'] if primero_activo else '-', 
            'circuito': primero_activo['circuito'] if primero_activo else '-',
            'duracion_segundos': getattr(orden, 'duracion_segundos', 0),
            'ultima_actualizacion_iso': orden.ultima_actualizacion.isoformat() if hasattr(orden, 'ultima_actualizacion') and orden.ultima_actualizacion else None,
        })
    
    return JsonResponse({'ordenes': data})

@require_http_methods(["GET","POST"])
def historial_finalizadas_view(request):
    def format_duration(seconds):
        if not seconds: return "00:00:00"
        seconds = int(seconds) 
        m, s = divmod(seconds, 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    # 1. --- LOGICA NUEVA: MAPA DE RETORNOS ---
    # Buscamos en TODAS las tablas (activas y finalizadas) quién está referenciando a una orden vieja
    # Queremos pares: (Referencia_Vieja, Orden_Nueva_Que_La_Creo)
    
    referencias = {} # Diccionario: Clave=OrdenVieja -> Valor=OrdenNueva
    
    # Buscamos en MCM
    refs_mcm = OrdenMCM.objects.exclude(orden_retorno_ref__isnull=True).exclude(orden_retorno_ref__exact='').values_list('orden_retorno_ref', 'numero_orden')
    # Buscamos en CHO, TPM, PREP...
    refs_cho = OrdenCHO.objects.exclude(orden_retorno_ref__isnull=True).exclude(orden_retorno_ref__exact='').values_list('orden_retorno_ref', 'numero_orden')
    refs_tpm = OrdenTPM.objects.exclude(orden_retorno_ref__isnull=True).exclude(orden_retorno_ref__exact='').values_list('orden_retorno_ref', 'numero_orden')
    refs_prep = OrdenPREP.objects.exclude(orden_retorno_ref__isnull=True).exclude(orden_retorno_ref__exact='').values_list('orden_retorno_ref', 'numero_orden')

    # Llenamos el diccionario maestro
    # Si la orden '1050' generó la '1055', el diccionario será: {'1050': '1055'}
    for ref, nueva in list(chain(refs_mcm, refs_cho, refs_tpm, refs_prep)):
        referencias[str(ref)] = str(nueva)

    # ------------------------------------------

    p_asignaciones = Prefetch('asignaciones', queryset=AsignacionUniversal.objects.all())
    
    qs_mcm = OrdenMCM.objects.filter(estado='Finalizada').select_related('molde', 'lider').prefetch_related(p_asignaciones).all()
    qs_cho = OrdenCHO.objects.filter(estado='Finalizada').select_related('molde', 'lider').prefetch_related(p_asignaciones).all()
    qs_tpm = OrdenTPM.objects.filter(estado='Finalizada').select_related('molde', 'lider').prefetch_related(p_asignaciones).all()
    qs_prep = OrdenPREP.objects.filter(estado='Finalizada').select_related('molde', 'lider').prefetch_related(p_asignaciones).all()

    todas = list(chain(qs_mcm, qs_cho, qs_tpm, qs_prep))
    todas.sort(key=lambda o: getattr(o, 'fecha_cierre', o.ultima_actualizacion) or o.fecha_creacion, reverse=True)
    
    datos = []
    for orden in todas:
        asignaciones = list(orden.asignaciones.all())
        fecha_inicio = orden.fecha_creacion
        fecha_fin = getattr(orden, 'fecha_cierre', orden.ultima_actualizacion)
        
        lista_detallada = []
        nombres_simples = []
        
        tiempo_activo_segundos = getattr(orden, 'duracion_segundos', 0)
        tiempo_real_segundos = 0
        if fecha_inicio and fecha_fin:
            tiempo_real_segundos = (fecha_fin - fecha_inicio).total_seconds()
            
        for item in asignaciones:
            detalles_str = ""
            if item.detalles_json:
                try:
                    import json
                    d_obj = json.loads(item.detalles_json)
                    nombres_defectos = [d.get('defecto') for d in d_obj if d.get('defecto')]
                    if nombres_defectos: detalles_str = ", ".join(nombres_defectos)
                except: pass
            if not detalles_str: detalles_str = item.defecto or '-'

            lista_detallada.append({
                'nombre': item.nombre_tecnico,
                'mesa': item.mesa or '-',       
                'cavidad': item.cavidad or '-',
                'circuito': item.circuito or '-',
                'defecto': detalles_str,
                'inicio': item.fecha_inicio.strftime('%H:%M') if item.fecha_inicio else '-',
                'fin': item.fecha_fin.strftime('%H:%M') if item.fecha_fin else 'Activo'
            })
            if item.nombre_tecnico: nombres_simples.append(item.nombre_tecnico)

        tecnico_tabla = ", ".join(list(set(nombres_simples)))
        
        # --- BUSCAMOS SI ESTA ORDEN GENERÓ UN RETORNO ---
        # Preguntamos: ¿El número de esta orden aparece como referencia en alguna otra?
        retorno_generado = referencias.get(str(orden.numero_orden), None)
        # ------------------------------------------------

        datos.append({
            'id': orden.id,
            'tipo': orden.tipo_mntn,
            'numero_orden': orden.numero_orden,
            'molde': orden.molde.nombre if orden.molde else 'N/A',
            'defecto': getattr(orden, 'defecto_sap', '-'),
            'lider': str(orden.lider) if orden.lider else 'Sin Asignar',
            'tecnico': tecnico_tabla,
            'lista_tecnicos': lista_detallada,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'tiempo_activo_fmt': format_duration(tiempo_activo_segundos),
            'tiempo_real_fmt': format_duration(tiempo_real_segundos),    
            'comentarios': orden.comentarios,

            # Datos propios (Si ella fue retorno)
            'orden_retorno_ref': getattr(orden, 'orden_retorno_ref', None),
            'motivo_retorno': getattr(orden, 'motivo_retorno', ''),
            'observaciones_retorno': getattr(orden, 'observaciones_retorno', ''),

            # --- NUEVO DATO: SI ELLA PROVOCÓ UN RETORNO ---
            'genero_retorno': retorno_generado  # Contendrá el número de la nueva orden (ej: "1055") o None
        })

    return render(request, 'Moldeo/ordenes_finalizadas.html', {'ordenes': datos})
@require_http_methods(["POST"])
@transaction.atomic
def api_actualizar_orden_view(request, orden_id):
    try:
        data = json.loads(request.body)
        tipo = data.get('tipo') 
        modelo_map = {'MCM': OrdenMCM, 'CHO': OrdenCHO, 'TPM': OrdenTPM, 'PREP': OrdenPREP}
        Modelo = modelo_map.get(tipo)
        
        if not Modelo: return JsonResponse({'message': 'Tipo inválido'}, status=400)

        # Usamos select_for_update para bloquear la fila mientras editamos
        orden = Modelo.objects.select_for_update().get(id=orden_id)

       
        if 'estado' in data:
            nuevo_estado_flujo = data['estado']
            ahora = timezone.now()
            
            # 1. DE ACTIVA -> PAUSADA (Detener reloj)
            if orden.estado == 'Activa' and nuevo_estado_flujo == 'Pausada':
                if orden.ultima_actualizacion:
                    segundos = (ahora - orden.ultima_actualizacion).total_seconds()
                    orden.duracion_segundos = (orden.duracion_segundos or 0) + int(segundos)
                orden.ultima_actualizacion = None 

            # 2. DE PAUSADA -> ACTIVA (Arrancar reloj)
            elif orden.estado == 'Pausada' and nuevo_estado_flujo == 'Activa':
                orden.ultima_actualizacion = ahora

            # 3. FINALIZAR
            elif nuevo_estado_flujo == 'Finalizada':
                if orden.estado == 'Activa' and orden.ultima_actualizacion:
                    segundos = (ahora - orden.ultima_actualizacion).total_seconds()
                    orden.duracion_segundos = (orden.duracion_segundos or 0) + int(segundos)
                
                orden.ultima_actualizacion = None
                if not orden.fecha_cierre: orden.fecha_cierre = ahora
                
                # Cerrar técnicos
                ct = ContentType.objects.get_for_model(orden)
                AsignacionUniversal.objects.filter(content_type=ct, object_id=orden.id, activo=True).update(activo=False, fecha_fin=ahora)

           
            orden.estado = nuevo_estado_flujo
            orden.save()

        # --- GUARDAR COMENTARIOS ---
        if 'comentarios' in data:
            orden.comentarios = data['comentarios']
            orden.save()

        return JsonResponse({'success': True})

    except Exception as e:
        print(f"ERROR API: {e}")
        return JsonResponse({'message': str(e)}, status=500)

def api_get_moldes(request):
    moldes = Moldes.objects.all().values('id_molde', 'numero_molde')
    data = [{'molde': m['numero_molde'], 'pk': m['id_molde']} for m in moldes]
    return JsonResponse(data, safe=False)

def exportar_ordenes_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Ordenes_General.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordenes General"

    headers = ['Tipo', 'ID', 'Orden', 'Status', 'Molde', 'Defecto SAP', 'Asignaciones (Técnico/Defecto)', 'Estado', 'Fecha', 'Comentarios']
    ws.append(headers)

    p_asignaciones = Prefetch('asignaciones', queryset=AsignacionUniversal.objects.all())
    
    qs_mcm = OrdenMCM.objects.select_related('molde').prefetch_related(p_asignaciones).all()
    qs_cho = OrdenCHO.objects.select_related('molde').prefetch_related(p_asignaciones).all()
    qs_tpm = OrdenTPM.objects.select_related('molde').prefetch_related(p_asignaciones).all()
    qs_prep = OrdenPREP.objects.select_related('molde').prefetch_related(p_asignaciones).all()

    todas = list(chain(qs_mcm, qs_cho, qs_tpm, qs_prep))
    todas.sort(key=attrgetter('fecha_creacion'), reverse=True)

    for orden in todas:
        if orden.molde:
            try: nombre_molde = orden.molde.numero_molde
            except: nombre_molde = "Ref Error"
        else: nombre_molde = "N/A"

        # Formatear asignaciones en una celda
        asignaciones_str = ""
        for item in orden.asignaciones.all():
            def_str = f" ({item.defecto})" if item.defecto else ""
            asignaciones_str += f"[{item.nombre_tecnico}{def_str} - {item.mesa or ''}], "

        ws.append([
            orden.tipo_mntn,            
            orden.id,
            orden.numero_orden,
            orden.status,              
            nombre_molde,
            getattr(orden, 'defecto_sap', '-'),
            asignaciones_str,
            orden.estado,
            timezone.localtime(orden.fecha_creacion).strftime('%Y-%m-%d %H:%M'),
            orden.comentarios
        ])

    wb.save(response)
    return response

@require_http_methods(["POST"])
def api_gestionar_tecnicos(request, orden_id):
    try:
        data = json.loads(request.body)
        accion = data.get('accion') 
        tipo = data.get('tipo')
        
        modelo_map = {'MCM': OrdenMCM, 'CHO': OrdenCHO, 'TPM': OrdenTPM, 'PREP': OrdenPREP}
        Modelo = modelo_map.get(tipo)
        if not Modelo: return JsonResponse({'error': 'Tipo inválido'}, status=400)
        
        try:
            orden = Modelo.objects.get(id=orden_id)
        except Modelo.DoesNotExist:
            return JsonResponse({'error': 'Orden no encontrada'}, status=404)

        if accion == 'agregar':
            nombre = data.get('nombre')
            mesa = data.get('mesa')
            
            # --- CORRECCIÓN: RECIBIR LISTA DE DETALLES ---
            detalles_lista = data.get('detalles', []) 
            
            # Si por alguna razón llega vacío, creamos uno genérico
            if not detalles_lista:
                detalles_lista = [{
                    'defecto': data.get('defecto', 'General'),
                    'cav': data.get('cavidad', '-'),
                    'circ': data.get('circuito', '-')
                }]

            # Serializamos a Texto para guardar en la BD
            detalles_json_str = json.dumps(detalles_lista)

            AsignacionUniversal.objects.create(
                content_object=orden,
                nombre_tecnico=nombre,
                mesa=mesa,
                detalles_json=detalles_json_str, # <--- Guardamos el JSON
                defecto="Ver detalles",          # Texto dummy para compatibilidad
                activo=True
            )

        elif accion == 'baja':
            item_id = data.get('item_id')
            asignacion = AsignacionUniversal.objects.get(id=item_id)
            asignacion.activo = False
            asignacion.fecha_fin = timezone.now()
            asignacion.save()

        elif accion == 'relevo':
            # ... (Tu lógica de relevo existente) ...
            id_saliente = data.get('item_id_saliente')
            nombre_entrante = data.get('nombre_entrante')
            
            saliente = AsignacionUniversal.objects.get(id=id_saliente)
            saliente.activo = False
            saliente.fecha_fin = timezone.now()
            saliente.save()

            # El nuevo hereda los detalles del anterior
            AsignacionUniversal.objects.create(
                content_object=orden,
                nombre_tecnico=nombre_entrante,
                mesa=saliente.mesa,
                detalles_json=saliente.detalles_json, # <--- Heredar JSON
                defecto=saliente.defecto,
                activo=True
            )

        return JsonResponse({'success': True})

    except Exception as e:
        print(f"ERROR API: {e}") # Ver error en terminal
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
def encontrar_encabezados(df_preview, keywords):
    """Busca la fila donde aparecen las palabras clave y retorna su índice."""
    for idx, row in df_preview.iterrows():
        fila_str = [str(val).strip() for val in row.values]
        if all(any(k.lower() in s.lower() for s in fila_str) for k in keywords):
            return idx
    return None

@login_required
def carga_masiva_view(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('archivo_excel') or request.FILES.get('archivo_sap')
        modelo_destino = request.POST.get('modelo_destino')

        if not excel_file:
            messages.error(request, "Por favor selecciona un archivo.")
            return render(request, 'Moldeo/subir_excel.html')

        try:
            registros_creados = 0
            
            # ==========================================
            # CASO A: ACTUALIZACIÓN SAP
            # ==========================================
            if modelo_destino == 'sap':
                if not excel_file.name.endswith('.xlsx'):
                    messages.error(request, 'Para SAP el archivo debe ser .xlsx')
                    return render(request, 'Moldeo/subir_excel.html')

                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
                
                header_row = 1
                headers = {}
                # Buscamos la fila de encabezados (buscando "Order")
                for r in range(1, 6):
                    row_vals = [str(cell.value).strip() for cell in ws[r] if cell.value]
                    if 'Order' in row_vals:
                        header_row = r
                        for cell in ws[r]:
                            if cell.value: headers[str(cell.value).strip()] = cell.column - 1
                        break
                
                if 'Order' not in headers:
                    messages.error(request, "Error SAP: No se encontró la columna 'Order'.")
                    return redirect('Moldeo:carga_masiva')

                # Índices de columnas
                idx_fecha = headers.get('Bas. start date') or headers.get('Basic start date')
                idx_hora = headers.get('Start time')
                idx_equipo = headers.get('Equipment')
                # --- NUEVO: Buscar columna de Status ---
                idx_status = headers.get('System status') or headers.get('SystemStatus')

                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    try:
                        order_val = row[headers['Order']]
                        if not order_val: continue
                        
                        # Procesar Fechas y Horas (Igual que antes)
                        fecha_val = None
                        if idx_fecha is not None and row[idx_fecha]:
                            raw = row[idx_fecha]
                            if hasattr(raw, 'date'): fecha_val = raw.date()
                            elif isinstance(raw, datetime): fecha_val = raw.date()

                        hora_val = None
                        if idx_hora is not None and row[idx_hora]:
                            raw = row[idx_hora]
                            if isinstance(raw, (datetime, time)):
                                hora_val = raw.time() if isinstance(raw, datetime) else raw
                            elif isinstance(raw, str):
                                try: hora_val = datetime.strptime(raw.strip(), "%H:%M:%S").time()
                                except: pass

                        equip = str(row[idx_equipo]).strip() if idx_equipo is not None and row[idx_equipo] else ''
                        
                        # --- NUEVO: Procesar Status ---
                        sys_status = ''
                        if idx_status is not None and row[idx_status]:
                            sys_status = str(row[idx_status]).strip()

                        # Guardar en BD
                        OrdenSAP.objects.update_or_create(
                            order=str(order_val),
                            defaults={
                                'description': row[headers.get('Description', -1)] if 'Description' in headers else '',
                                'work_center': row[headers.get('Work center', -1)] if 'Work center' in headers else '',
                                'equipment': equip,
                                'fecha_inicio': fecha_val,
                                'hora_inicio': hora_val,
                                'system_status': sys_status # <--- Guardamos el status aquí
                            }
                        )
                        registros_creados += 1
                    except: continue

                messages.success(request, f'SAP Actualizado: {registros_creados} órdenes procesadas.')
                return redirect('Moldeo:panel_principal')

            # ==========================================
            # CASO B: CATÁLOGOS (Pandas)
            # ==========================================
            else:
                df = None
                
                # Leer archivo
                if excel_file.name.endswith('.csv'):
                    try: df = pd.read_csv(excel_file, encoding='utf-8')
                    except: df = pd.read_csv(excel_file, encoding='latin-1')
                else:
                    xls = pd.ExcelFile(excel_file)
                    # Palabras clave para detectar encabezados
                    keywords = []
                    if modelo_destino == 'moldes': keywords = ['MOLDE', 'MAQUINA']
                    elif modelo_destino == 'defectos': keywords = ['DEFECTOS']
                    elif modelo_destino == 'actividades_tpm': keywords = ['ACTIVIDADES TPM', 'ZONAS']
                    elif modelo_destino == 'estatus_ordenes': keywords = ['STATUS', 'SYSTEM STATUS']
                    else: keywords = ['nombre'] # Default
                    
                    for sheet in xls.sheet_names:
                        preview = pd.read_excel(excel_file, sheet_name=sheet, nrows=10, header=None)
                        idx = encontrar_encabezados(preview, keywords)
                        if idx is not None:
                            df = pd.read_excel(excel_file, sheet_name=sheet, header=idx)
                            break
                    
                    if df is None: df = pd.read_excel(excel_file)

                # Limpieza de nombres de columnas
                df.columns = df.columns.astype(str).str.strip()
                col_names = df.columns.tolist()

                # --- PROCESO MOLDES ---
                if modelo_destino == 'moldes':
                    col_maestro = 'MOLDE' if 'MOLDE' in col_names else next((c for c in col_names if c.upper()=='MOLDE'), None)
                    col_maquina = 'MAQUINA' if 'MAQUINA' in col_names else next((c for c in col_names if c.upper()=='MAQUINA'), None)
                    
                    col_detalle = 'Molde' 
                    if 'Molde' not in col_names and 'Molde.1' in col_names: col_detalle = 'Molde.1'
                    elif 'Molde' not in col_names and col_maestro: 
                        col_detalle = next((c for c in col_names if 'Molde' in c and c != col_maestro), None)

                    col_parte = 'Numeros de Parte' if 'Numeros de Parte' in col_names else next((c for c in col_names if 'Parte' in c), None)

                    if not col_maestro:
                        messages.error(request, f"Error: No se encontró la columna maestra 'MOLDE'.")
                        return redirect('Moldeo:carga_masiva')

                    # PASO 1: MAESTRO
                    moldes_ok = 0
                    df_master = df.dropna(subset=[col_maestro])
                    
                    for _, row in df_master.iterrows():
                        nm = str(row[col_maestro]).strip()
                        if not nm or nm.lower() == 'nan' or nm == 'EOAT': continue
                        
                        instancia_maq = None
                        if col_maquina and str(row[col_maquina]).lower() != 'nan':
                            maq_nombre = str(row[col_maquina]).strip()
                            if maq_nombre:
                                instancia_maq, _ = Maquinas.objects.get_or_create(nombre=maq_nombre)
                        
                        proy = None
                        if 'PROYECTO' in row and str(row['PROYECTO']).lower() != 'nan':
                            proy = str(row['PROYECTO']).strip()

                        Moldes.objects.update_or_create(
                            nombre=nm,
                            defaults={'maquina': instancia_maq, 'proyecto': proy, 'activo': True}
                        )
                        moldes_ok += 1

                    # PASO 2: DETALLES
                    partes_ok = 0
                    if col_detalle and col_parte:
                        df_det = df.dropna(subset=[col_parte])

                        for _, row in df_det.iterrows():
                            ref_molde = str(row[col_detalle]).strip()
                            num_parte = str(row[col_parte]).strip()

                            if not ref_molde or ref_molde.lower() == 'nan': continue
                            if not num_parte or num_parte.lower() == 'nan': continue

                            molde_obj = Moldes.objects.filter(nombre=ref_molde).first()

                            if molde_obj:
                                if 'Molde SAP' in row and str(row['Molde SAP']).lower() != 'nan':
                                    molde_obj.molde_sap = str(row['Molde SAP']).strip()
                                
                                if 'Cavidades' in row:
                                    try:
                                        c = int(row['Cavidades'])
                                        if c > 0: molde_obj.cavidades = c
                                    except: pass
                                
                                molde_obj.save()

                                NumerosParte.objects.get_or_create(
                                    numero_parte=num_parte,
                                    molde=molde_obj
                                )
                                partes_ok += 1

                    messages.success(request, f"Proceso OK: {moldes_ok} moldes actualizados. {partes_ok} partes vinculadas.")

                # --- PROCESO ACTIVIDADES TPM (NUEVO) ---
                elif modelo_destino == 'actividades_tpm':
                    actividades_creadas = 0
                    actividadesPREP_creadas = 0
                    zonas_creadas = 0
                    
                    # Detectar columnas clave
                    col_act = next((c for c in col_names if 'ACTIVIDADES TPM' in c.upper()), None)
                    col_zona = next((c for c in col_names if 'ZONAS' in c.upper()), None)
                    col_subzona = next((c for c in col_names if 'SUBZONAS' in c.upper()), None)
                    col_req = next((c for c in col_names if 'COMENTARIOS' in c.upper() or 'REQUIERE' in c.upper()), None) # Columna J en tu excel

                    # 1. Cargar Actividades (Columna A)
                    if col_act:
                        df_act = df.dropna(subset=[col_act])
                        for _, row in df_act.iterrows():
                            nombre = str(row[col_act]).strip()
                            if nombre and nombre.lower() != 'nan':
                                ActividadTPM.objects.get_or_create(nombre=nombre)
                                actividades_creadas += 1

                    # 2. Cargar Zonas y Subzonas (Columnas G, H)
                    if col_zona and col_subzona:
                        df_zonas = df.dropna(subset=[col_zona])
                        for _, row in df_zonas.iterrows():
                            z_nombre = str(row[col_zona]).strip()
                            s_nombre = str(row[col_subzona]).strip()
                            
                            # Logica para "Requiere detalles"
                            requiere = False
                            if col_req and str(row[col_req]).lower() != 'nan':
                                comm = str(row[col_req]).lower()
                                if 'requiere' in comm or 'si' in comm:
                                    requiere = True

                            if z_nombre and z_nombre.lower() != 'nan':
                                zona_obj, _ = ZonaTPM.objects.get_or_create(nombre=z_nombre)
                                
                                if s_nombre and s_nombre.lower() != 'nan':
                                    SubZonaTPM.objects.update_or_create(
                                        zona=zona_obj,
                                        nombre=s_nombre,
                                        defaults={'requiere_detalles': requiere}
                                    )
                                    zonas_creadas += 1
                    # Detectar columnas clave
                    col_act = next((c for c in col_names if 'ACTIVIDADES PREP' in c.upper()), None)
                        

                            # 1. Cargar Actividades 
                    if col_act:
                        df_act = df.dropna(subset=[col_act])
                        for _, row in df_act.iterrows():
                            nombre = str(row[col_act]).strip()
                            if nombre and nombre.lower() != 'nan':
                                ActividadPREP.objects.get_or_create(nombre=nombre)
                                actividadesPREP_creadas += 1
                    messages.success(request, f"TPM Cargado: {actividades_creadas} actividades, PREP Cargado: {actividades_creadas} actividades, PREP Cargado: {actividadesPREP_creadas} actividades y {zonas_creadas} subzonas.")
                       
                            
                            
                            
                elif modelo_destino == 'estatus_ordenes':
                    estatus_creados = 0
                    
                    # 1. Buscar columnas exactas del Excel
                    col_sys = next((c for c in col_names if c.strip().upper() == 'SYSTEM STATUS'), None)
                    col_desc = next((c for c in col_names if c.strip().upper() == 'STATUS'), None)

                    if col_sys and col_desc:
                        for _, row in df.iterrows():
                            # Limpieza de datos
                            val_status = str(row[col_sys]).strip()      # Ej: REL PRT MANC...
                            val_desc = str(row[col_desc]).strip()       # Ej: ABIERTA
                            
                            if val_status and val_status.lower() != 'nan':
                                # Guardar en BD usando tus campos nuevos
                                EstatusOrden.objects.update_or_create(
                                    status=val_status,  # Campo del modelo = Valor del Excel
                                    defaults={'descripcion': val_desc}
                                )
                                estatus_creados += 1
                        messages.success(request, f"¡Éxito! {estatus_creados} estatus cargados.")
                    else:
                        messages.error(request, f"Error: No se encontraron las columnas 'SYSTEM STATUS' y 'STATUS'.")

                # --- OTROS CATÁLOGOS ---
                elif modelo_destino == 'defectos':
                    col = next((c for c in col_names if 'DEFECTOS' in c.upper()), None)
                    if col:
                        for _, row in df.iterrows():
                            nom = str(row[col]).strip()
                            if nom and nom.lower() != 'nan':
                                Defectos.objects.get_or_create(nombre_defecto=nom, defaults={'activo': True})
                                registros_creados += 1
                        messages.success(request, f"{registros_creados} defectos cargados.")

                elif modelo_destino == 'maquinas':
                    col = next((c for c in col_names if 'MAQUINA' in c.upper()), None)
                    if col:
                        for _, row in df.iterrows():
                            nm = str(row[col]).strip()
                            if nm and nm.lower() != 'nan':
                                Maquinas.objects.get_or_create(nombre=nm)
                                registros_creados += 1
                        messages.success(request, f"{registros_creados} máquinas cargadas.")

                elif modelo_destino in ['tecnicos', 'lideres']:
                    col = 'nombre' if 'nombre' in col_names else col_names[0]
                    Modelo = Moldmakers if modelo_destino == 'tecnicos' else Lideres
                    for _, row in df.iterrows():
                        nm = str(row[col]).strip()
                        if nm and nm.lower() != 'nan':
                            Modelo.objects.get_or_create(nombre=nm)
                            registros_creados += 1
                    messages.success(request, f"{registros_creados} personas cargadas.")

            return redirect('Moldeo:carga_masiva')

        except Exception as e:
            messages.error(request, f"Error crítico: {str(e)}")
            return render(request, 'Moldeo/subir_excel.html')

    return render(request, 'Moldeo/subir_excel.html')
   
def lista_sap_view(request):
    # --- 1. Filtros y Búsqueda (Tu lógica estándar) ---
    busqueda = request.GET.get('q', '')
    fecha_filtro = request.GET.get('fecha', '')
 
    ids_registrados = list(chain(
        OrdenMCM.objects.values_list('numero_orden', flat=True),
        OrdenCHO.objects.values_list('numero_orden', flat=True),
        OrdenTPM.objects.values_list('numero_orden', flat=True),
        OrdenPREP.objects.values_list('numero_orden', flat=True)
    ))
    
    ordenes = OrdenSAP.objects.exclude(order__in=ids_registrados)

    if busqueda:
        ordenes = ordenes.filter(
            Q(order__icontains=busqueda) |
            Q(work_center__icontains=busqueda) | # Busca por Molde
            Q(description__icontains=busqueda)
        )
    
    if fecha_filtro:
        ordenes = ordenes.filter(fecha_inicio=fecha_filtro)

  
    ordenes = ordenes.order_by('-fecha_inicio')

   
    moldes_con_maquina = Moldes.objects.select_related('maquina').filter(maquina__isnull=False)

    diccionario_maquinas = {}
    for m in moldes_con_maquina:
        nombre_molde = str(m.nombre).strip()
      
        nombre_maquina = m.maquina.nombre
        diccionario_maquinas[nombre_molde] = nombre_maquina

    # C. Pegamos la etiqueta a cada orden
    lista_final = []
    for orden in ordenes:
        # ¿Qué molde pide esta orden?
        molde_solicitado = str(orden.work_center).strip() if orden.work_center else ""
        
        # Buscamos en el diccionario
        maquina_encontrada = diccionario_maquinas.get(molde_solicitado)
        
        # Guardamos el dato en una variable temporal dentro del objeto
        if maquina_encontrada:
            orden.maquina_relacionada = maquina_encontrada
        else:
            orden.maquina_relacionada = None # No tiene máquina asignada
            
        lista_final.append(orden)

    context = {
        'ordenes': lista_final,
        'busqueda': busqueda,
        'fecha': fecha_filtro
    }
    return render(request, 'Moldeo/lista_sap.html', context)
def imprimir_formato_view(request, orden_id, tipo):
    # Buscamos la orden según el tipo
    orden = None
    if tipo == 'MCM':
        orden = OrdenMCM.objects.get(id=orden_id)
    elif tipo == 'TPM':
        orden = OrdenTPM.objects.get(id=orden_id)
    # ... agregar otros tipos si es necesario ...

    context = {
        'orden': orden,
        # Pasamos la fecha actual para la impresión si se requiere
        'fecha_impresion': timezone.now()
    }
    return render(request, 'Moldeo/imprimir_formato.html', context)
def api_buscar_orden_info(request):
    numero = request.GET.get('numero')
    if not numero:
        return JsonResponse({'success': False, 'message': 'Falta número'})
    
    orden = OrdenMCM.objects.filter(numero_orden=numero).last()
    
    if orden:
        lista_asignaciones = []
        for asig in orden.asignaciones.all():
            detalles = []
            if asig.detalles_json:
                import json
                try:
                    detalles = json.loads(asig.detalles_json)
                except: pass
            
            # LÓGICA DE SEPARACIÓN (Parsing)
            raw_name = asig.nombre_tecnico
            nombre_real = raw_name
            lider_real = ""
            
            # Si el nombre tiene el formato "Nombre (L: Lider)"
            if "(L:" in raw_name:
                try:
                    parts = raw_name.split("(L:")
                    nombre_real = parts[0].strip() # "Juan Perez"
                    lider_real = parts[1].replace(")", "").strip() # "Pedro"
                except:
                    pass

            lista_asignaciones.append({
                'nombre': nombre_real,
                'lider': lider_real, # Enviamos el líder limpio al JS
                'mesa': asig.mesa,
                'detalles': detalles
            })

        return JsonResponse({
            'success': True,
            'molde': str(orden.molde),
            'defecto': orden.defecto_sap,
            'comentarios': orden.comentarios or '',
            'motivo_retorno': orden.motivo_retorno or '',
            'observaciones_retorno': orden.observaciones_retorno or '',
            'asignaciones': lista_asignaciones
        })
    else:
        return JsonResponse({'success': False, 'message': 'Orden no encontrada'})
    
@require_http_methods(["GET"])
def api_filtrar_ordenes_mcm(request):
    """
    Retorna una lista simple de órdenes FINALIZADAS para el autocompletado.
    """
    q = request.GET.get('q', '')
    
  
    if not q: 
        return JsonResponse([], safe=False)
    
    qs = OrdenMCM.objects.filter(
        numero_orden__icontains=q,
        estado='Finalizada' 
    ).order_by('-fecha_creacion')[:10]
    
    data = []
    for o in qs:
        nombre_molde = o.molde.nombre if o.molde else 'N/A'
        data.append({
            'numero': o.numero_orden,
            'molde': nombre_molde,
            'defecto': o.defecto_sap
        })
        
    return JsonResponse(data, safe=False)
def historial_molde_api(request):
    molde_nombre = request.GET.get('molde')
    if not molde_nombre:
        return JsonResponse({'success': False, 'message': 'Falta nombre de molde'})

    try:
        # Buscar el objeto molde
        molde_obj = Moldes.objects.filter(nombre=molde_nombre).first() # Ajusta 'numero_molde' o 'nombre' según tu modelo
        if not molde_obj:
             # Intento buscar por nombre exacto si usas otro campo
             molde_obj = Moldes.objects.filter(nombre=molde_nombre).first()

        if not molde_obj:
             return JsonResponse({'success': True, 'historial': []}) # No existe molde, retorna vacío

        # Buscar en las 3 tablas (MCM, CHO, TPM)
        # Filtramos por molde y estado finalizado
        mcm = OrdenMCM.objects.filter(molde=molde_obj, estado='Finalizada').order_by('-fecha_cierre')[:3]
        cho = OrdenCHO.objects.filter(molde=molde_obj, estado='Finalizada').order_by('-fecha_cierre')[:3]
        tpm = OrdenTPM.objects.filter(molde=molde_obj, estado='Finalizada').order_by('-fecha_cierre')[:3]

        # Unir y ordenar
        todos = sorted(
            list(chain(mcm, cho, tpm)),
            key=lambda x: x.fecha_cierre if x.fecha_cierre else x.fecha_creacion,
            reverse=True
        )[:3] # Tomamos solo las 3 más recientes de todas

        data = []
        for o in todos:
            # Obtener defecto según el tipo de modelo
            defecto_txt = getattr(o, 'defecto_sap', '') or getattr(o, 'comentarios', '')
            
            data.append({
                'orden': o.numero_orden,
                'tipo': o.tipo_mntn, # Asegúrate que tu modelo tenga este campo o hardcodealo
                'fecha': o.fecha_cierre.strftime('%d/%m/%Y') if o.fecha_cierre else '---',
                'defecto': defecto_txt[:30] + '...' if len(defecto_txt) > 30 else defecto_txt,
                'lider': o.lider.username if o.lider else '---'
            })

        return JsonResponse({'success': True, 'historial': data})

    except Exception as e:
        print(f"Error API Historial: {e}")
        return JsonResponse({'success': False, 'message': str(e)})
    
def detalle_orden_historial_api(request):
    orden_id = request.GET.get('orden')
    tipo = request.GET.get('tipo') # MCM, CHO, TPM

    if not orden_id or not tipo:
        return JsonResponse({'success': False, 'message': 'Faltan parámetros.'})

    try:
        orden = None
        # Buscar en el modelo correcto
        if tipo == 'MCM':
            orden = OrdenMCM.objects.filter(numero_orden=orden_id).first()
        elif tipo == 'CHO':
            orden = OrdenCHO.objects.filter(numero_orden=orden_id).first()
        elif tipo == 'TPM':
            orden = OrdenTPM.objects.filter(numero_orden=orden_id).first()
        
        if not orden:
            return JsonResponse({'success': False, 'message': 'Orden no encontrada.'})

        # Buscar técnicos asignados (Relación Genérica)
        tecnicos_qs = AsignacionUniversal.objects.filter(
            object_id=orden.pk, 
            content_type__model=f'orden{tipo.lower()}' # ej: ordenmcm
        )
        
        tecnicos_list = []
        for t in tecnicos_qs:
            tecnicos_list.append({
                'nombre': t.nombre_tecnico,
                'mesa': t.mesa
            })

        data = {
            'success': True,
            'orden': orden.numero_orden,
            'fecha': orden.fecha_cierre.strftime('%d/%m/%Y %H:%M') if orden.fecha_cierre else '---',
            'tipo': orden.tipo_mntn,
            'lider': orden.lider.username if orden.lider else 'N/A',
            'comentarios': orden.comentarios,
            'tecnicos': tecnicos_list
        }
        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})
def api_ordenes_pendientes(request):
    molde_id = request.GET.get('molde_id')
    print(f"--- DEBUG API --- Buscando órdenes para Molde ID: {molde_id}")

    if not molde_id:
        return JsonResponse({'ordenes': []})

    try:
        # 1. Recopilar IDs de órdenes que YA existen (Convertidos a String para comparar seguro)
        # Usamos 'values_list' con 'flat=True' para obtener listas simples ['1001', '1002']
        ids_cho = set(str(x) for x in OrdenCHO.objects.values_list('numero_orden', flat=True))
        ids_tpm = set(str(x) for x in OrdenTPM.objects.values_list('numero_orden', flat=True))
        ids_mcm = set(str(x) for x in OrdenMCM.objects.values_list('numero_orden', flat=True))
        
        # Unimos todos los sets en uno solo para búsqueda rápida
        ids_ocupados = ids_cho.union(ids_tpm).union(ids_mcm)
        print(f"--- DEBUG API --- Total órdenes ya ocupadas en el sistema: {len(ids_ocupados)}")

        # 2. Buscar en OrdenSAP
        # IMPORTANTE: Aquí asumo que OrdenSAP tiene una relación ForeignKey llamada 'molde'.
        # Si OrdenSAP se relaciona por texto (work_center), avísame para cambiar esta línea.
        #qs_sap = OrdenSAP.objects.filter(molde_id=molde_id)
        try:
            molde_obj = Moldes.objects.get(pk=molde_id)
            # Asumiendo que el nombre del molde es igual al work_center en SAP
            qs_sap = OrdenSAP.objects.filter(work_center=molde_obj.nombre) 
        except Moldes.DoesNotExist:
            qs_sap = []
        print(f"--- DEBUG API --- Órdenes SAP encontradas para este molde (Total bruto): {qs_sap.count()}")

        data = []
        for orden in qs_sap:
            # Convertimos el número de orden SAP a string
            num_sap = str(orden.order) 
            
            # Verificamos si NO está en la lista de ocupados
            if num_sap not in ids_ocupados:
                data.append({
                    'numero': orden.order,       
                    'descripcion': orden.description or 'Sin descripción'
                })
        
        print(f"--- DEBUG API --- Órdenes disponibles retornadas al frontend: {len(data)}")
        return JsonResponse({'ordenes': data})

    except Exception as e:
        print(f"--- DEBUG API ERROR --- Ocurrió un error: {str(e)}")
        return JsonResponse({'ordenes': []})
@login_required
def panel_lider_view(request):
    # 1. OBTENER ÓRDENES EN CURSO (Piso)
    ordenes_cho = OrdenCHO.objects.filter(estado__in=['Activa', 'Pausada']).order_by('-fecha_creacion')
    ordenes_mcm = OrdenMCM.objects.filter(estado__in=['Activa', 'Pausada']).order_by('-fecha_creacion') # <--- Agregado MCM
    ordenes_tpm = OrdenTPM.objects.filter(estado__in=['Activa', 'Pausada']).order_by('-fecha_creacion')
    
    # Corregido: Suma de los tres tipos
    total_en_curso = ordenes_cho.count() + ordenes_mcm.count() + ordenes_tpm.count()
    
    # Lista negra (Órdenes ya en piso para excluirlas del backlog SAP)
    ordenes_ocupadas = set(ordenes_cho.values_list('numero_orden', flat=True))
    ordenes_ocupadas.update(set(ordenes_mcm.values_list('numero_orden', flat=True))) # <--- Agregado MCM
    ordenes_ocupadas.update(set(ordenes_tpm.values_list('numero_orden', flat=True)))

    # 2. OBTENER TODAS LAS DE SAP (Backlog completo - Excluyendo las que ya están en piso)
    pendientes_sap_qs = OrdenSAP.objects.exclude(order__in=ordenes_ocupadas).order_by('order')

    # 3. CÁLCULO DE MÉTRICAS (Clasificación por Estatus)
    # Primero obtenemos los códigos de estatus de nuestro catálogo
    codigos_abierta = EstatusOrden.objects.filter(descripcion__icontains='ABIERTA').values_list('status', flat=True)
    
    codigos_mal = EstatusOrden.objects.filter(
        Q(descripcion__icontains='MAL') | 
        Q(descripcion__icontains='INCOMPLETO') |
        Q(descripcion__icontains='ERROR')
    ).values_list('status', flat=True)
    
    # Cerradas son las que dicen CERRADA pero NO son MAL/INCOMPLETO
    codigos_cerrada = EstatusOrden.objects.filter(descripcion__icontains='CERRADA').exclude(status__in=codigos_mal).values_list('status', flat=True)

    # Contamos directo en la BD (Count es rápido)
    count_abiertas = pendientes_sap_qs.filter(system_status__in=codigos_abierta).count()
    count_problemas = pendientes_sap_qs.filter(system_status__in=codigos_mal).count()
    count_cerradas = pendientes_sap_qs.filter(system_status__in=codigos_cerrada).count()

    # 4. TÉCNICOS
    tecnicos = Moldmakers.objects.filter(activo=True).order_by('nombre')
    
    # Unificamos TODAS las activas para la columna derecha (CHO + MCM + TPM)
    from itertools import chain
    # Se usa sorted para ordenarlas por fecha (la más reciente arriba), opcional
    activas_todas = list(chain(ordenes_cho, ordenes_mcm, ordenes_tpm))
    activas_todas.sort(key=lambda x: x.fecha_creacion, reverse=True)

    context = {
        'pendientes': pendientes_sap_qs, 
        'activas': activas_todas,
        'tecnicos': tecnicos,
        'metrics': {
            'en_curso': total_en_curso,
            'abiertas': count_abiertas,      # Trabajo Real
            'problemas': count_problemas,    # Mal Cerradas / Incompletas
            'cerradas': count_cerradas       # Cerradas OK
        }
    }
    return render(request, 'Moldeo/panel_lider.html', context)
@login_required
def api_panel_lider_data(request):
    """Retorna el HTML actualizado para las listas del panel"""
    
    # 1. Órdenes en Curso (Misma lógica que la vista principal)
    ordenes_cho = OrdenCHO.objects.filter(estado__in=['Activa', 'Pausada'])
    ordenes_mcm = OrdenMCM.objects.filter(estado__in=['Activa', 'Pausada'])
    ordenes_tpm = OrdenTPM.objects.filter(estado__in=['Activa', 'Pausada'])
    
    # Unificar y ordenar
    from itertools import chain
    activas_todas = list(chain(ordenes_cho, ordenes_mcm, ordenes_tpm))
    activas_todas.sort(key=lambda x: x.fecha_creacion, reverse=True)

    # 2. Órdenes Pendientes (Misma lógica)
    ordenes_ocupadas = set(ordenes_cho.values_list('numero_orden', flat=True))
    ordenes_ocupadas.update(set(ordenes_mcm.values_list('numero_orden', flat=True)))
    ordenes_ocupadas.update(set(ordenes_tpm.values_list('numero_orden', flat=True)))

    pendientes_sap = OrdenSAP.objects.exclude(order__in=ordenes_ocupadas).order_by('order')

    # Renderizamos SOLO las partes parciales
    # (Necesitaremos crear un template parcial o usar render_to_string)
    from django.template.loader import render_to_string
    
    html_pendientes = render_to_string('Moldeo/partials/lista_pendientes.html', {'pendientes': pendientes_sap})
    html_activas = render_to_string('Moldeo/partials/lista_activas.html', {'activas': activas_todas})
    
    return JsonResponse({
        'html_pendientes': html_pendientes,
        'html_activas': html_activas
    })
@login_required
def panel_kiosco_view(request):
    """Vista principal del Kiosco"""
    tecnicos = Moldmakers.objects.filter(activo=True).order_by('nombre')
    
    # Órdenes activas para mostrar en la lista general
    ordenes_cho = OrdenCHO.objects.filter(estado__in=['Activa', 'Pausada'])
    ordenes_mcm = OrdenMCM.objects.filter(estado__in=['Activa', 'Pausada'])
    ordenes_tpm = OrdenTPM.objects.filter(estado__in=['Activa', 'Pausada'])
    
    from itertools import chain
    activas = list(chain(ordenes_cho, ordenes_mcm, ordenes_tpm))
    activas.sort(key=lambda x: x.fecha_creacion, reverse=True)

    context = {
        'tecnicos': tecnicos,
        'activas': activas,
    }
    return render(request, 'Moldeo/panel_kiosco.html', context)

@login_required
def api_kiosco_login(request):
    """Valida credenciales del técnico sin loguear en Django"""
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        tech_id = data.get('tech_id')
        password = data.get('password')

        try:
            tecnico = Moldmakers.objects.get(id=tech_id)
            # Validación simple (puedes mejorarla si usas hash)
            if tecnico.password == password or password == 'MASTER123': # Backdoor opcional
                
                # Buscar órdenes asignadas a ESTE técnico
                mis_ordenes = []
                # Buscar en Asignaciones Universales donde nombre_tecnico contenga su nombre
                # (Es una búsqueda aproximada porque guardamos strings, idealmente sería FK)
                asignaciones = AsignacionUniversal.objects.filter(
                    nombre_tecnico__icontains=tecnico.nombre,
                    activo=True,
                    content_type__model__in=['ordencho', 'ordenmcm', 'ordentpm']
                )
                
                for asig in asignaciones:
                    orden = asig.content_object
                    if orden.estado in ['Activa', 'Pausada']:
                        mis_ordenes.append({
                            'id': orden.id,
                            'numero': orden.numero_orden,
                            'tipo': orden.tipo_mntn, # Asegúrate que el modelo tenga esta propiedad o string
                            'molde': str(orden.molde),
                            'mesa': asig.mesa
                        })

                return JsonResponse({'success': True, 'nombre': tecnico.nombre, 'ordenes': mis_ordenes})
            else:
                return JsonResponse({'success': False, 'message': 'Contraseña incorrecta'})
        except Moldmakers.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Técnico no encontrado'})
    return JsonResponse({'success': False})

@login_required
def api_kiosco_sumarse(request):
    """Permite a un técnico validado sumarse a una orden"""
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        tech_id = data.get('tech_id')
        orden_numero = data.get('orden_numero')
        tipo_orden = data.get('tipo_orden') # CHO, MCM, TPM

        try:
            tecnico = Moldmakers.objects.get(id=tech_id)
            
            # Buscar la orden
            orden = None
            if tipo_orden == 'CHO':
                orden = OrdenCHO.objects.filter(numero_orden=orden_numero).first()
            elif tipo_orden == 'MCM':
                orden = OrdenMCM.objects.filter(numero_orden=orden_numero).first()
            elif tipo_orden == 'TPM':
                orden = OrdenTPM.objects.filter(numero_orden=orden_numero).first()

            if orden:
                # Crear asignación
                AsignacionUniversal.objects.create(
                    content_object=orden,
                    nombre_tecnico=tecnico.nombre,
                    mesa='-', # Mesa genérica al sumarse
                    activo=True
                )
                return JsonResponse({'success': True, 'message': f'Te has sumado a la orden {orden_numero}'})
            else:
                return JsonResponse({'success': False, 'message': 'Orden no encontrada'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False})